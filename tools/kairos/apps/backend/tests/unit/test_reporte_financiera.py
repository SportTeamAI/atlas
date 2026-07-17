"""Quién entra y quién NO entra al reporte a Financiera.

El archivo que se le manda a Financiera es de RECARGOS Y EXTRAS, no de la nómina
completa: si a alguien no hay que pagarle nada de más, su fila sobra. Esta regla toca
plata, así que va con test.
"""

from jornada.application.reporte_excel import COL_CATEGORIAS, construir_excel_area, tiene_recargo

from datetime import date


def _fila(nombre: str, **cats) -> dict:
    return {"cedula": "1234", "nombre": nombre, "cargo": "Analista", "cats": cats, "observaciones": ""}


def test_sin_ningun_recargo_no_va_al_reporte():
    # Solo trabajó ordinaria diurna: no hay recargo ni extra que pagar.
    assert tiene_recargo(_fila("Sin recargos", ORD_DIUR_REG=44.0)) is False


def test_sin_horas_del_todo_no_va():
    assert tiene_recargo(_fila("Vacío")) is False
    assert tiene_recargo({"nombre": "Sin cats"}) is False


def test_cualquier_recargo_o_extra_si_va():
    # Cada columna del Excel, por sí sola, basta para que la persona entre.
    for _col, _nombre, cat, _cm in COL_CATEGORIAS:
        assert tiene_recargo(_fila("Con recargo", **{cat: 1.0})) is True, cat


def test_recargo_en_cero_no_cuenta():
    # Estar en la categoría con 0.0 h no es tener recargo.
    assert tiene_recargo(_fila("Cero", ORD_NOCT_REG=0.0)) is False


def test_el_excel_solo_trae_a_quien_tiene_recargo():
    filas = [
        _fila("Ana Con Nocturno", ORD_NOCT_REG=5.0),
        _fila("Beto Solo Ordinaria", ORD_DIUR_REG=44.0),   # no debe aparecer
        _fila("Cami Con Extra", EXT_DIUR_REG=2.0),
    ]
    xlsx = construir_excel_area(
        area="Pruebas", periodo="NM1QJulio26", fecha_corte=date(2026, 7, 8),
        fecha_inicio=date(2026, 7, 1), fecha_fin=date(2026, 7, 8), filas=filas,
    )
    from io import BytesIO

    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(xlsx)).active
    nombres = {c.value for row in ws.iter_rows(min_col=2, max_col=2) for c in row if c.value}
    assert "Ana Con Nocturno" in nombres
    assert "Cami Con Extra" in nombres
    assert "Beto Solo Ordinaria" not in nombres, "quien no tiene recargo no debe ir a Financiera"
