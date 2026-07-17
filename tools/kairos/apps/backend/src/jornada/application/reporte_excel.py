"""Genera el Excel "Recargos - <ÁREA>" replicando la plantilla del cliente.

Incluye: logo de VirtualSoft en A1, encabezados D4..K4 EXACTOS con su comentario,
bordes solo hasta donde hay nombres, y en C3 el rango del período ("21 de mayo
al 1 de junio").
"""

from __future__ import annotations

import io
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_LOGO = Path(__file__).with_name("logo_financiera.png")
_MESES = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio",
          "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

# (letra columna, encabezado EXACTO con espacios, categoría de horas, comentario).
COL_CATEGORIAS = [
    ("D", "RECARGO NOCTURNO ", "ORD_NOCT_REG",
     "Es el trabajo que se da entre las 9pm a 12pm de Lunes a Sábado."),
    ("E", "DOMINGO  O FESTIVO TRABAJADO DIURNO   ", "ORD_DIUR_DESC",
     "Horas laboradas un Domingo o día festivo diurno entre las 6am y 9pm."),
    ("F", "DOMINGO  O FESTIVO TRABAJADO NOCTURNO        ", "ORD_NOCT_DESC",
     "Horas laboradas un Domingo o día festivo entre las 9pm a 12pm"),
    ("G", "HORAS EXTRAS DIURNAS               ", "EXT_DIUR_REG",
     "Son las horas adicionales al turno que se dan entre las 6am y 9pm de Lunes a Sábado."),
    ("H", "HORAS EXTRAS NOCTURNA ", "EXT_NOCT_REG",
     "Son las horas adicionales al turno que se dan entre las 7 pm a 12pm de Lunes a Sábado."),
    ("I", "HORAS EXTRAS FESTIVA O DOMINICAL DIURNA         ", "EXT_DIUR_DESC",
     "Horas adicionales al turno trabajadas un Domingo o día festivo entre las 6am y 9pm "),
    ("J", "HORAS EXTRAS  FESTIVA O DOMINICAL NOCTURNA             ", "EXT_NOCT_DESC",
     "Horas adicionales al turno trabajadas un Domingo o día festivo entre las 9pm y las 12pm"),
]
ANCHOS = {"A": 13.7, "B": 31.3, "C": 35.5, "D": 13.2, "E": 13.8, "F": 14.5,
          "G": 15.3, "H": 13.7, "I": 18.7, "J": 18.5, "K": 28.5}

_GRIS = PatternFill("solid", fgColor="D9D9D9")
_THIN = Side(style="thin", color="000000")
_BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _hdr(cell, texto, size=8):
    cell.value = texto
    cell.font = Font(bold=True, size=size)
    cell.fill = _GRIS
    cell.alignment = _CENTER
    cell.border = _BORDE


def _fmt_rango(ini: date, fin: date) -> str:
    return f"{ini.day} de {_MESES[ini.month]} al {fin.day} de {_MESES[fin.month]}"


def tiene_recargo(fila: dict) -> bool:
    """¿Esta persona tiene algo que reportar a Financiera? Las columnas del Excel son SOLO
    recargos y extras (la ordinaria diurna no va): si todas están en cero, no hay nada que
    pagar de más y la fila sobra."""
    cats = fila.get("cats") or {}
    return any(cats.get(cat) for _col, _nombre, cat, _cm in COL_CATEGORIAS)


def construir_excel_area(*, area: str, periodo: str, fecha_corte: date,
                         fecha_inicio: date, fecha_fin: date, filas: list[dict]) -> bytes:
    """`filas`: [{cedula, nombre, cargo, cats: {CATEGORIA: horas}, observaciones}].

    Quien NO tenga ningún recargo/extra NO se incluye: el archivo es de recargos, no de
    la nómina completa.
    """
    filas = [f for f in filas if tiene_recargo(f)]
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for col, w in ANCHOS.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 53
    ws.row_dimensions[4].height = 61

    # Bordes en TODA la cabecera A1:K3 (#4).
    for row in range(1, 4):
        for c in range(1, 12):
            ws.cell(row=row, column=c).border = _BORDE

    # Logo en A1 (#15): más pequeño, centrado y anclado a la celda (queda fijo).
    # #9 Si falta Pillow o el logo, se sigue sin logo (no romper la descarga).
    ws.merge_cells("A1:B1")
    if _LOGO.exists():
      try:
        img = XLImage(str(_LOGO))
        img.width, img.height = 96, 38  # más pequeño
        # Anclar dentro de A1 con un pequeño margen para que quede centrado y fijo.
        from openpyxl.drawing.spreadsheet_drawing import (
            AnchorMarker,
            OneCellAnchor,
        )
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        marker = AnchorMarker(col=0, colOff=pixels_to_EMU(24), row=0, rowOff=pixels_to_EMU(16))
        img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(96), pixels_to_EMU(38)))
        ws.add_image(img)
      except Exception:  # noqa: BLE001 — sin Pillow/logo se sigue sin logo (#9)
        pass

    # Título y área.
    ws.merge_cells("C1:K1")
    t = ws["C1"]; t.value = "REPORTE DE TIEMPO - EXTRAS"
    t.font = Font(bold=True, size=15); t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("G2:K3")
    a = ws["G2"]; a.value = f"Area: {area}"
    a.font = Font(bold=True, size=15); a.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:B2"); _hdr(ws["A2"], "FECHA DE ENVÍO DE LA NOVEDAD:", size=10)
    ws.merge_cells("C2:D2"); ws["C2"].value = fecha_corte.isoformat(); ws["C2"].alignment = _CENTER; ws["C2"].border = _BORDE
    ws.merge_cells("A3:B3"); _hdr(ws["A3"], "PERIODO A LIQUIDAR:", size=10)
    # #19: C3 con el rango del período ("21 de mayo al 1 de junio").
    ws.merge_cells("C3:D3"); ws["C3"].value = _fmt_rango(fecha_inicio, fecha_fin); ws["C3"].alignment = _CENTER; ws["C3"].border = _BORDE

    # Encabezados fila 4 (exactos) + comentarios (#18).
    _hdr(ws["A4"], "CEDULA"); _hdr(ws["B4"], "NOMBRE"); _hdr(ws["C4"], "CARGO")
    for col, nombre, _cat, comentario in COL_CATEGORIAS:
        c = ws[f"{col}4"]; _hdr(c, nombre)
        c.comment = Comment(comentario, "Talento Humano")
    _hdr(ws["K4"], "OBSERVACIONES", size=6)

    # Filas de empleados (#17 rellena los recargos). Bordes SOLO hasta donde hay
    # nombres (#16).
    r = 5
    for f in filas:
        ws.cell(row=r, column=1, value=f["cedula"])
        ws.cell(row=r, column=2, value=f["nombre"])
        ws.cell(row=r, column=3, value=f.get("cargo") or "")
        for idx, (_col, _nombre, cat, _cm) in enumerate(COL_CATEGORIAS, start=4):
            h = f["cats"].get(cat, 0)
            ws.cell(row=r, column=idx, value=round(h, 2) if h else None)
        ws.cell(row=r, column=11, value=f.get("observaciones") or "")
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = _BORDE
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_zip_periodo(carpeta: str, archivos: list[tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for nombre, contenido in archivos:
            z.writestr(f"{carpeta}/{nombre}", contenido)
    return buf.getvalue()
